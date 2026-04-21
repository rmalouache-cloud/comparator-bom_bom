import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
from PIL import Image

st.set_page_config(page_title="BOM Comparator", layout="wide")

# Logo TV
col_logo1, col_logo2, col_logo3 = st.columns([1, 2, 1])
with col_logo2:
    st.image("https://cdn-icons-png.flaticon.com/512/1792/1792772.png", width=100)  # Logo TV
    # Ou utilisez votre propre fichier local:
    # st.image("tv_logo.png", width=100)

st.title("📊 BOM Comparison Tool")

old_file = st.file_uploader("📂 Upload OLD BOM", type=["xlsx"])
new_file = st.file_uploader("📂 Upload NEW BOM", type=["xlsx"])

# =========================
# SAFE FUNCTION
# =========================
def safe_join(x):
    if not isinstance(x, list):
        return ""
    return ", ".join(str(i) for i in x if pd.notna(i))

def extract_components_by_type(df, bom_type):
    """
    Extrait les composants CKD ou SKD selon la ligne de démarcation
    CKD: de "ASS'Y - MAIN BOARD（CKD）" jusqu'à "BARCODE LABEL"
    SKD: tout le reste
    """
    if bom_type == "CKD":
        start_idx = None
        for idx, desc in enumerate(df['Description']):
            if 'ASS\'Y - MAIN BOARD（CKD）' in str(desc).upper() or 'ASSY - MAIN BOARD（CKD）' in str(desc).upper():
                start_idx = idx
                break
        
        end_idx = None
        for idx, desc in enumerate(df['Description']):
            if 'BARCODE LABEL' in str(desc).upper():
                end_idx = idx
                break
        
        if start_idx is not None and end_idx is not None:
            return df.iloc[start_idx:end_idx+1].copy()
        elif start_idx is not None:
            return df.iloc[start_idx:].copy()
        else:
            return pd.DataFrame()
    else:  # SKD
        start_idx = None
        for idx, desc in enumerate(df['Description']):
            if 'ASS\'Y - MAIN BOARD（CKD）' in str(desc).upper() or 'ASSY - MAIN BOARD（CKD）' in str(desc).upper():
                start_idx = idx
                break
        
        end_idx = None
        for idx, desc in enumerate(df['Description']):
            if 'BARCODE LABEL' in str(desc).upper():
                end_idx = idx
                break
        
        if start_idx is not None and end_idx is not None:
            before_ckd = df.iloc[:start_idx].copy()
            after_ckd = df.iloc[end_idx+1:].copy()
            return pd.concat([before_ckd, after_ckd], ignore_index=True)
        elif start_idx is not None:
            return df.iloc[:start_idx].copy()
        else:
            return df.copy()

def run_comparison(old_df, new_df, bom_type_name):
    """Exécute la comparaison entre deux DataFrames"""
    
    if old_df.empty and new_df.empty:
        return pd.DataFrame()
    
    cols = ["PN", "Description", "bom_qty", "BOM text"]
    
    # S'assurer que les colonnes existent
    old = old_df[cols].copy() if not old_df.empty else pd.DataFrame(columns=cols)
    new = new_df[cols].copy() if not new_df.empty else pd.DataFrame(columns=cols)
    
    if old.empty and new.empty:
        return pd.DataFrame()
    
    if not old.empty:
        old.rename(columns={"BOM text": "Position"}, inplace=True)
    if not new.empty:
        new.rename(columns={"BOM text": "Position"}, inplace=True)
    
    # Clean data
    for df in [old, new]:
        if not df.empty:
            df["PN"] = df["PN"].astype(str).str.strip().str.upper()
            df["Description"] = df["Description"].astype(str).str.strip().str.upper()
            df["Position"] = df["Position"].astype(str).str.strip().str.upper()
            
            df["bom_qty"] = (
                df["bom_qty"]
                .astype(str)
                .str.replace(",", ".", regex=False)
            )
            df["bom_qty"] = pd.to_numeric(df["bom_qty"], errors="coerce").fillna(0)
    
    # Group by PN only
    if not old.empty:
        old = old.groupby(["PN"], as_index=False).agg({
            "Description": "first",
            "bom_qty": "sum",
            "Position": list
        })
    else:
        old = pd.DataFrame(columns=["PN", "Description", "bom_qty", "Position"])
    
    if not new.empty:
        new = new.groupby(["PN"], as_index=False).agg({
            "Description": "first",
            "bom_qty": "sum",
            "Position": list
        })
    else:
        new = pd.DataFrame(columns=["PN", "Description", "bom_qty", "Position"])
    
    # Merge on PN
    df = old.merge(
        new,
        on="PN",
        how="outer",
        suffixes=("_old", "_new"),
        indicator=True
    )
    
    # Logic
    def get_status(row):
        if row["_merge"] == "left_only":
            return "Missing in BOM2"
        
        if row["_merge"] == "right_only":
            return "Missing in BOM1"
        
        qty_old = row["bom_qty_old"]
        qty_new = row["bom_qty_new"]
        
        pos_old = set(row["Position_old"]) if isinstance(row["Position_old"], list) else set()
        pos_new = set(row["Position_new"]) if isinstance(row["Position_new"], list) else set()
        
        if qty_old != qty_new:
            return "Qty diff"
        
        if pos_old != pos_new:
            return "Position diff"
        
        return "Conform"
    
    df["Status"] = df.apply(get_status, axis=1)
    
    # Build result
    result = []
    
    for _, row in df.iterrows():
        pos_old = row["Position_old"] if isinstance(row["Position_old"], list) else []
        pos_new = row["Position_new"] if isinstance(row["Position_new"], list) else []
        
        result.append({
            "Type": bom_type_name,
            "PN": row["PN"] if row["_merge"] != "right_only" else "",
            "Desc OLD": row.get("Description_old", ""),
            "Qty OLD": row.get("bom_qty_old", 0),
            "Pos OLD": safe_join(pos_old),
            "PN NEW": row["PN"] if row["_merge"] != "left_only" else "",
            "Desc NEW": row.get("Description_new", ""),
            "Qty NEW": row.get("bom_qty_new", 0),
            "Pos NEW": safe_join(pos_new),
            "Status": row["Status"]
        })
    
    return pd.DataFrame(result)

# =========================
# OPTIONS DE COMPARAISON
# =========================
st.markdown("---")
st.subheader("🔧 Options de comparaison")

col1, col2, col3 = st.columns(3)

with col1:
    compare_ckd = st.checkbox("📟 Comparer CKD", value=True)
with col2:
    compare_skd = st.checkbox("🔌 Comparer SKD", value=True)
with col3:
    compare_both = st.checkbox("🔄 Comparer CKD + SKD ensemble", value=False)

start = st.button("🚀 Démarrer la comparaison", use_container_width=True)

if start:
    
    if old_file is None or new_file is None:
        st.error("❌ Veuillez uploader les deux fichiers")
        st.stop()
    
    if not compare_ckd and not compare_skd and not compare_both:
        st.error("❌ Veuillez sélectionner au moins une option de comparaison")
        st.stop()
    
    # Read files
    old = pd.read_excel(old_file)
    new = pd.read_excel(new_file)
    
    old.columns = old.columns.str.strip()
    new.columns = new.columns.str.strip()
    
    all_results = []
    
    # Comparaison CKD seule
    if compare_ckd:
        with st.spinner("🔍 Comparaison CKD en cours... (📟 carte mère)"):
            old_ckd = extract_components_by_type(old, "CKD")
            new_ckd = extract_components_by_type(new, "CKD")
            
            if old_ckd.empty and new_ckd.empty:
                st.warning("⚠️ Aucun composant CKD trouvé")
            else:
                st.info(f"📟 CKD: {len(old_ckd)} composants dans OLD, {len(new_ckd)} dans NEW")
                result_ckd = run_comparison(old_ckd, new_ckd, "CKD")
                if not result_ckd.empty:
                    all_results.append(result_ckd)
    
    # Comparaison SKD seule
    if compare_skd:
        with st.spinner("🔍 Comparaison SKD en cours... (🔌 câbles)"):
            old_skd = extract_components_by_type(old, "SKD")
            new_skd = extract_components_by_type(new, "SKD")
            
            if old_skd.empty and new_skd.empty:
                st.warning("⚠️ Aucun composant SKD trouvé")
            else:
                st.info(f"🔌 SKD: {len(old_skd)} composants dans OLD, {len(new_skd)} dans NEW")
                result_skd = run_comparison(old_skd, new_skd, "SKD")
                if not result_skd.empty:
                    all_results.append(result_skd)
    
    # Comparaison CKD + SKD ensemble (fusionnés)
    if compare_both:
        with st.spinner("🔄 Comparaison complète CKD + SKD en cours..."):
            old_ckd_full = extract_components_by_type(old, "CKD")
            old_skd_full = extract_components_by_type(old, "SKD")
            old_full = pd.concat([old_ckd_full, old_skd_full], ignore_index=True)
            
            new_ckd_full = extract_components_by_type(new, "CKD")
            new_skd_full = extract_components_by_type(new, "SKD")
            new_full = pd.concat([new_ckd_full, new_skd_full], ignore_index=True)
            
            st.info(f"🔄 Comparaison complète: {len(old_full)} composants dans OLD, {len(new_full)} dans NEW")
            result_both = run_comparison(old_full, new_full, "COMPLET")
            if not result_both.empty:
                all_results.append(result_both)
    
    # Combiner tous les résultats
    if all_results:
        final_result = pd.concat(all_results, ignore_index=True)
        
        # Fix Streamlit crash
        for col in final_result.columns:
            final_result[col] = final_result[col].astype(str)
        
        st.success(f"✅ Comparaison terminée! {len(final_result)} lignes trouvées")
        st.dataframe(final_result, use_container_width=True)
        
        # Export Excel with colors
        output = io.BytesIO()
        final_result.to_excel(output, index=False)
        output.seek(0)
        
        wb = load_workbook(output)
        ws = wb.active
        
        colors = {
            "Conform": "C6EFCE",
            "Missing in BOM1": "FFC7CE",
            "Missing in BOM2": "FFC7CE",
            "Qty diff": "FFEB9C",
            "Position diff": "BDD7EE"
        }
        
        status_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Status":
                status_col = col
                break
        
        if status_col:
            for row in range(2, ws.max_row + 1):
                status = ws.cell(row=row, column=status_col).value
                
                for key, color in colors.items():
                    if status == key:
                        fill = PatternFill(start_color=color, fill_type="solid")
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = fill
        
        final_file = io.BytesIO()
        wb.save(final_file)
        final_file.seek(0)
        
        st.download_button(
            "📥 Télécharger Excel",
            final_file,
            "BOM_comparison.xlsx"
        )
    else:
        st.error("❌ Aucun résultat à afficher")
